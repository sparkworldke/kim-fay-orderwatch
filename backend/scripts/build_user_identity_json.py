#!/usr/bin/env python3
"""Build JSON seed data from Active staff Excel + HODs CSV (name match)."""

from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[2]
EXCEL = ROOT / "excels" / "Active staff July 2026- HQ.xlsx"
HODS = ROOT / "excels" / "Kimfay Employees - HODs.csv"
KP_CUSTOMERS = ROOT / "excels" / "KP-Customers 20260805 Agusst.csv"
OUT = ROOT / "backend" / "database" / "seeders" / "data"

# employee_number → Acumatica rep_code when they differ
REP_ALIASES = {
    "P317": "YVON",
    "P460": "C967",
    "P483": "C1262",
}


def norm_name(s: object) -> str:
    if not s:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"[^a-z0-9\s']", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def name_tokens(s: object) -> set[str]:
    stop = {"of", "and", "the", "de", "van"}
    return {t for t in norm_name(s).split() if t not in stop and len(t) > 1}


def load_permanent(wb) -> list[dict]:
    ws = wb["Permanent staff"]
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = list(row) + [None] * 5
        emp, name, dept, div, desig = cells[:5]
        if not emp or not name:
            continue
        emp_s = str(emp).strip().upper()
        if "EMPLOYEE" in emp_s:
            continue
        rows.append(
            {
                "employee_number": emp_s,
                "name": str(name).strip(),
                "name_key": norm_name(name),
                "department": str(dept).strip() if dept else None,
                "division": str(div).strip() if div else None,
                "designation": str(desig).strip() if desig else None,
                "sheet": "permanent",
            }
        )
    return rows


def load_stc(wb) -> list[dict]:
    ws = wb["STC"]
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        cells = list(row)
        if len(cells) < 5:
            continue
        emp, name, dept, div = cells[1], cells[2], cells[3], cells[4]
        if not emp or not name:
            continue
        emp_s = str(emp).strip().upper()
        if emp_s == "EMPLOYEE NUMBER":
            continue
        rows.append(
            {
                "employee_number": emp_s,
                "name": str(name).strip(),
                "name_key": norm_name(name),
                "department": str(dept).strip() if dept else None,
                "division": str(div).strip() if div else None,
                "designation": None,
                "sheet": "stc",
            }
        )
    return rows


def load_hods() -> list[dict]:
    hods: list[dict] = []
    with HODS.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if len(row) < 6:
                continue
            reports_to = (row[0] or "").strip().upper() or None
            emp = (row[1] or "").strip().upper()
            name = (row[2] or "").strip()
            if not emp or not name:
                continue
            hods.append(
                {
                    "reports_to": reports_to,
                    "employee_number": emp,
                    "name": name,
                    "name_key": norm_name(name),
                    "department": (row[3] or "").strip() or None,
                    "division": (row[4] or "").strip() or None,
                    "designation": (row[5] or "").strip() or None,
                    "email": (row[6] or "").strip().lower() or None,
                    "segment": (row[7] or "").strip() if len(row) > 7 else None,
                }
            )
    return hods


def find_staff_match(h: dict, perm: list[dict], by_name: dict[str, list[dict]]) -> dict | None:
    hits = by_name.get(h["name_key"], [])
    if hits:
        return hits[0]

    htoks = name_tokens(h["name"])
    best = None
    best_score = 0.0
    for p in perm:
        ptoks = name_tokens(p["name"])
        if not htoks or not ptoks:
            continue
        inter = len(htoks & ptoks)
        score = inter / max(len(htoks), len(ptoks))
        if inter >= 2 and score >= 0.5 and score > best_score:
            best_score = score
            best = p
    return best


def load_kp_rep_by_name() -> dict[str, dict]:
    """Sales Rep display name → rep_code from KP customers export."""
    out: dict[str, dict] = {}
    if not KP_CUSTOMERS.exists():
        return out
    with KP_CUSTOMERS.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            code = (row.get("Rep Code") or "").strip().upper()
            name = (row.get("Sales Rep") or "").strip()
            if not code or not name or code in ("LITIGATION", "BADDEBTPRO"):
                continue
            key = norm_name(name)
            emp = code if re.match(r"^[PC]\d+", code) else None
            for emp_no, alias in REP_ALIASES.items():
                if alias == code:
                    emp = emp_no
                    break
            out[key] = {"name": name, "rep_code": code, "employee_number": emp}
    return out


def write_kp_sales_rep_json(rep_by_name: dict[str, dict]) -> None:
    reps = [
        {
            "match_names": [v["name"]],
            "rep_code": v["rep_code"],
            "employee_number": v["employee_number"],
            "source": "kp_customers_20260805",
        }
        for v in rep_by_name.values()
    ]
    aliases = [
        {
            "match_names": [name],
            "employee_number": emp,
            "rep_code": code,
        }
        for emp, code in REP_ALIASES.items()
        for name in {
            "P317": ["Yvonne Achieng Otieno", "Yvonne Achieng"],
            "P460": ["Berna Piwang Abondo", "Berna Piwang"],
            "P483": ["Mercyline Kemunto Moranga", "Mercyline Kemunto"],
        }.get(emp, [])
    ]
    (OUT / "kp-sales-rep-codes-2026-08.json").write_text(
        json.dumps(
            {
                "description": "Rep codes from KP-Customers CSV + known aliases.",
                "reps": reps,
                "aliases": aliases,
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def main() -> None:
    if not EXCEL.exists():
        raise SystemExit(f"Missing {EXCEL}")
    if not HODS.exists():
        raise SystemExit(f"Missing {HODS}")

    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    perm = load_permanent(wb)
    stc = load_stc(wb)
    hods = load_hods()

    # Prefer permanent staff for matches; fall back to STC roster.
    all_staff_for_match = perm + stc
    by_name: dict[str, list[dict]] = {}
    for p in all_staff_for_match:
        by_name.setdefault(p["name_key"], []).append(p)

    alignments: list[dict] = []
    unmatched_hod: list[dict] = []

    for h in hods:
        p = find_staff_match(h, all_staff_for_match, by_name)
        if not p:
            unmatched_hod.append(h)
            continue
        alignments.append(
            {
                "name": p["name"],
                "match_name": h["name"],
                "name_key": p["name_key"],
                # Prefer Active staff employee number for users.employee_number
                "employee_number": p["employee_number"],
                "hod_employee_number": h["employee_number"],
                "reports_to_employee_number": h["reports_to"],
                "department": h["department"] or p["department"],
                "division": h["division"] or p["division"],
                "designation": h["designation"] or p["designation"],
                "email": h["email"],
                "segment": h["segment"],
                "numbers_match": p["employee_number"] == h["employee_number"],
                "match_type": "exact_name" if p["name_key"] == h["name_key"] else "fuzzy_name",
            }
        )

    OUT.mkdir(parents=True, exist_ok=True)

    staff_payload = {
        "generated_from": EXCEL.name,
        "permanent_count": len(perm),
        "stc_count": len(stc),
        "staff": perm + stc,
    }
    (OUT / "active-staff-2026-07.json").write_text(
        json.dumps(staff_payload, indent=2), encoding="utf-8"
    )

    hods_payload = {
        "generated_from": HODS.name,
        "count": len(hods),
        "hods": hods,
    }
    (OUT / "hods-2026.json").write_text(json.dumps(hods_payload, indent=2), encoding="utf-8")

    # User corrections: match existing users by name → set employee_number from staff.
    # Include all permanent staff as lookup by name for non-HOD users too.
    user_corrections = []
    seen_keys: set[str] = set()
    for a in alignments:
        key = a["employee_number"]
        if key in seen_keys:
            continue
        seen_keys.add(key)
        user_corrections.append(
            {
                "match_names": list({a["name"], a["match_name"]}),
                "employee_number": a["employee_number"],
                "hod_employee_number": a["hod_employee_number"],
                "reports_to_employee_number": a["reports_to_employee_number"],
                "department": a["department"],
                "division": a["division"],
                "designation": a["designation"],
                "email": a["email"],
                "segment": a["segment"],
                "source": "hod_matched_to_active_staff",
            }
        )

    for p in perm:
        if p["employee_number"] in seen_keys:
            continue
        seen_keys.add(p["employee_number"])
        user_corrections.append(
            {
                "match_names": [p["name"]],
                "employee_number": p["employee_number"],
                "hod_employee_number": None,
                "reports_to_employee_number": None,
                "department": p["department"],
                "division": p["division"],
                "designation": p["designation"],
                "email": None,
                "segment": None,
                "source": "active_staff_only",
            }
        )

    align_payload = {
        "description": (
            "Name matches between Active staff July 2026 HQ and Kimfay Employees HODs. "
            "users.employee_number should be set from Active staff (employee_number). "
            "hod_employee_number is the org-chart code from HODs CSV when present."
        ),
        "generated_from": [EXCEL.name, HODS.name],
        "matched_hod_count": len(alignments),
        "unmatched_hod_count": len(unmatched_hod),
        "unmatched_hods": unmatched_hod,
        "number_mismatches": [a for a in alignments if not a["numbers_match"]],
        "alignments": alignments,
        "user_corrections": user_corrections,
    }
    # Enrich corrections with rep_code from KP customers + aliases
    rep_by_name = load_kp_rep_by_name()
    for c in user_corrections:
        emp = c.get("employee_number")
        if emp in REP_ALIASES:
            c["rep_code"] = REP_ALIASES[emp]
            c["rep_code_source"] = "known_alias"
            continue
        for mn in c.get("match_names") or []:
            key = norm_name(mn)
            if key in rep_by_name:
                c["rep_code"] = rep_by_name[key]["rep_code"]
                c["rep_code_source"] = "kp_customers_sales_rep_name"
                break
            toks = set(name_tokens(mn))
            for rk, rv in rep_by_name.items():
                if len(toks & set(rk.split())) >= 2:
                    c["rep_code"] = rv["rep_code"]
                    c["rep_code_source"] = "kp_customers_sales_rep_name_fuzzy"
                    break
            if c.get("rep_code"):
                break

    align_payload["user_corrections"] = user_corrections
    (OUT / "user-identity-alignments-2026-08.json").write_text(
        json.dumps(align_payload, indent=2), encoding="utf-8"
    )

    seed_users = [
        {
            "match_names": c["match_names"],
            "employee_number": c["employee_number"],
            "rep_code": c.get("rep_code"),
            "rep_code_source": c.get("rep_code_source"),
            "department": c.get("department"),
            "division": c.get("division"),
            "designation": c.get("designation"),
            "email": c.get("email"),
            "reports_to_employee_number": c.get("reports_to_employee_number"),
            "source": c.get("source"),
        }
        for c in user_corrections
    ]
    (OUT / "user-identity-seed-2026-08.json").write_text(
        json.dumps(
            {
                "description": align_payload["description"],
                "generated_from": align_payload["generated_from"],
                "users": seed_users,
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    write_kp_sales_rep_json(rep_by_name)

    print(f"Permanent staff: {len(perm)}")
    print(f"STC: {len(stc)}")
    print(f"HODs: {len(hods)}")
    print(f"HOD matched to staff: {len(alignments)}")
    print(f"HOD unmatched: {len(unmatched_hod)}")
    print(f"Number mismatches: {sum(1 for a in alignments if not a['numbers_match'])}")
    print(f"Seed users with rep_code: {sum(1 for u in seed_users if u.get('rep_code'))}")
    for a in alignments:
        if not a["numbers_match"]:
            print(
                f"  MISMATCH {a['match_name']}: HOD={a['hod_employee_number']} STAFF={a['employee_number']}"
            )
    for h in unmatched_hod:
        print(f"  UNMATCHED HOD {h['employee_number']} {h['name']}")
    print(f"Wrote JSON under {OUT}")


if __name__ == "__main__":
    main()
